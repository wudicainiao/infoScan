import os
import json
import queue
import openpyxl as ws
from gevent import socket
from lib.config import define




def creat_xlsx(q_results):
    if os.path.exists(define.filename) == False:
        s = 0
        wb = ws.Workbook()
        ws1 = wb.active
        if os.path.exists('out\\') == False:
            os.mkdir('out')
        word=['host','ip','title','ports_open','cidr','asn','org','addr','isp','cdn','url','waf(True为存在waf)']
        for i in word:
            s = s + 1
            ws1.cell(row =1,column = s,value = i)
        wb.save(define.filename)
        #wb.close()
        q_results.put("[*]创建文件成功 %s"%define.filename)
    else:
        q_results.put("[*]文件已存在 文件为:%s 请稍后运行程序"%define.filename)

def write_xlsx(q_targets, q_results):
    #first open file
    wb = ws.load_workbook(define.filename)
    q_results.put("[*]结果正在写入文件")
    while True:
        try:
            target = q_targets.get(timeout=0.2)
        except queue.Empty:
            break
    #second write row in memory
        sheet1 = wb['Sheet']
        num = sheet1.max_row
        sheet1.cell(row = num+1,column = 1,value = target['host'])
        sheet1.cell(row = num+1,column = 2,value = target['ip'])
        sheet1.cell(row = num+1,column = 3,value = str(target['title']))
        sheet1.cell(row = num+1,column = 4,value = str(target['ports_open']))
        sheet1.cell(row = num+1,column = 5,value = target['cidr'])
        sheet1.cell(row = num+1,column = 6,value = target['asn'])
        sheet1.cell(row = num+1,column = 7,value = target['org'])
        sheet1.cell(row = num+1,column = 8,value = target['addr'])
        sheet1.cell(row = num+1,column = 9,value = target['isp'])
        sheet1.cell(row = num+1,column = 10,value = target['cdn'])
        sheet1.cell(row = num+1,column = 11,value = str(target['url']))
        sheet1.cell(row = num+1,column = 12,value = str(target['waf']))
    #thrid write and close
    wb.save(define.filename)
    wb.close()
    q_results.put("[*]结果写入完成 : %s"%define.filename)

def is_port_open(host, port):
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(3.0)
        if s.connect_ex((host, int(port))) == 0:
            return True
        else:
            return False
    except Exception as e:
        return False
    finally:
        try:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_LINGER, struct.pack('ii', 1, 0))
            s.close()
        except Exception as e:
            pass

def scan_given_ports(host, ports_open, port='80'):
    if port in define.ports:
        for p in define.ports:
            if is_port_open(host, p):
                ports_open.add(p)
    else:
        if is_port_open(host, port):
            ports_open.add(port)
        for p in define.ports:
            if is_port_open(host, p):
                ports_open.add(p)
